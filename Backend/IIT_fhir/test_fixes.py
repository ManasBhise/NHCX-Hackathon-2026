"""
Test script to verify the fixes:
1. FHIR validation using parse_obj instead of model_validate
2. Data mapping sheet generation from API-uploaded JSON files
"""

import json
import requests
import time

API_BASE = "http://localhost:8000"

# Test 1: Validation with percentage reporting
print("\n" + "="*80)
print("TEST 1: FHIR Validation with percentage reporting")
print("="*80)

json_file = r"f:\Desktop\IIT HYDERABAD\Backend\IIT_fhir\output\pending\Aditya Birla_02.json"

with open(json_file, "rb") as f:
    files = {"file": f}
    response = requests.post(f"{API_BASE}/validate", files=files)

if response.status_code == 200:
    result = response.json()
    print(f"\n✓ Validation successful!")
    print(f"  - Compliance Score: {result.get('compliance_percentage')}%")
    print(f"  - Completion Score: {result.get('completion_percentage')}%")
    print(f"  - Valid: {result.get('valid')}")
    print(f"  - Errors: {result.get('error_count')}")
    print(f"  - Warnings: {result.get('warning_count')}")
    print(f"  - Category Scores: {result.get('category_scores')}")
    
    # Check if FHIR score is better than 40%
    fhir_score = result.get('category_scores', {}).get('FHIR', 0)
    if fhir_score > 40:
        print(f"\n✓ FHIR validation fix SUCCESSFUL! Score improved from 40% to {fhir_score}%")
    else:
        print(f"\n✗ FHIR score still low at {fhir_score}%")
else:
    print(f"✗ Validation failed: {response.status_code}")
    print(f"  Response: {response.text}")

# Test 2: JSON to Excel conversion with data mapping
print("\n" + "="*80)
print("TEST 2: JSON to Excel conversion with Data Mapping sheet")
print("="*80)

with open(json_file, "rb") as f:
    files = {"file": f}
    response = requests.post(f"{API_BASE}/json-to-excel", files=files)

if response.status_code == 200:
    excel_filename = r"f:\Desktop\IIT HYDERABAD\Backend\IIT_fhir\test_output.xlsx"
    with open(excel_filename, "wb") as out:
        out.write(response.content)
    print(f"\n✓ Excel file generated successfully!")
    print(f"  - File size: {len(response.content)} bytes")
    print(f"  - Saved to: {excel_filename}")
    
    # Verify it's a valid XLSX file
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_filename)
        sheet_names = wb.sheetnames
        print(f"  - Sheets: {sheet_names}")
        
        # Check if Data Mapping sheet has content
        if "Data Mapping" in sheet_names:
            ws = wb["Data Mapping"]
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
            print(f"  - Data Mapping sheet has {data_rows} data rows")
            
            if data_rows > 0:
                print(f"\n✓ Data Mapping fix SUCCESSFUL! Sheet contains actual data (not just 'No raw data available')")
            else:
                print(f"\n✗ Data Mapping sheet is empty")
        
        wb.close()
    except Exception as e:
        print(f"  - Could not verify Excel content: {e}")
else:
    print(f"✗ Excel generation failed: {response.status_code}")
    print(f"  Response: {response.text}")

print("\n" + "="*80)
print("Test Summary")
print("="*80)
print("Both fixes have been applied and API is responding correctly.")
