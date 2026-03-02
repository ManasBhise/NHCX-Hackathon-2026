"""
Generate Excel mapping and output files from FHIR JSON bundles.

Creates a comprehensive Excel file showing:
- Mapping rules applied (from mapping.yaml)
- Extracted Organization data
- Extracted InsurancePlan data
- Benefits and Coverage details
- Claim Exclusions
"""

import json
import yaml
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

logger = None


def setup_styles():
    """Define reusable styles for Excel sheets."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    return {
        "header_fill": header_fill,
        "header_font": header_font,
        "subheader_fill": subheader_fill,
        "subheader_font": subheader_font,
        "border": border
    }


def format_header(ws, row, columns, styles):
    """Format a header row."""
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = col_title
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 25


def format_subheader(ws, row, title, styles, merge_cols=4):
    """Format a subheader cell and merge columns."""
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.fill = styles["subheader_fill"]
    cell.font = styles["subheader_font"]
    cell.border = styles["border"]
    ws.merge_cells(f"A{row}:{get_column_letter(merge_cols)}{row}")
    ws.row_dimensions[row].height = 20


def _reconstruct_raw_data_from_bundle(bundle):
    """
    Reconstruct raw extraction data from FHIR bundle.
    
    When raw LLM file is not available (e.g., API upload), extract key fields
    from the FHIR bundle to show in the Data Mapping sheet.
    
    Returns a dict with fields that match raw LLM extraction format.
    """
    raw_data = {}
    
    try:
        # Find Organization and InsurancePlan resources
        org = None
        plan = None
        
        for entry in bundle.get("entry", []):
            resource = entry.get("resource", {})
            if resource.get("resourceType") == "Organization":
                org = resource
            elif resource.get("resourceType") == "InsurancePlan":
                plan = resource
        
        # Extract from Organization
        if org:
            raw_data["organization"] = org.get("name", "")
            
            # Extract identifiers
            if org.get("identifier"):
                raw_data["insurer_id"] = org["identifier"][0].get("value", "")
            
            # Extract telecom
            raw_data["telecom"] = {}
            for telecom in org.get("telecom", []):
                if telecom.get("system") == "phone":
                    raw_data["telecom"]["phone"] = telecom.get("value", "")
                elif telecom.get("system") == "email":
                    raw_data["telecom"]["email"] = telecom.get("value", "")
        
        # Extract from InsurancePlan
        if plan:
            # UIN/identifier
            if plan.get("identifier"):
                raw_data["uin"] = plan["identifier"][0].get("value", "")
            
            raw_data["plan_name"] = plan.get("name", "")
            raw_data["plan_type"] = plan.get("type", [{}])[0].get("coding", [{}])[0].get("display", "")
            raw_data["status"] = plan.get("status", "")
            
            # Extract benefits
            benefits = []
            for coverage_area in plan.get("coverageArea", []):
                benefit = {
                    "name": coverage_area.get("display", ""),
                    "coverage_percentage": "100"  # Default
                }
                benefits.append(benefit)
            
            if benefits:
                raw_data["benefits"] = benefits
            
            # Extract exclusions
            exclusions = []
            for extension in plan.get("extension", []):
                if extension.get("url") == "https://nrces.in/ndhm/fhir/r4/StructureDefinition/Claim-Exclusion":
                    for sub_ext in extension.get("extension", []):
                        if sub_ext.get("url") == "category":
                            exclusion = {
                                "name": sub_ext.get("valueCodeableConcept", {}).get("text", "")
                            }
                            exclusions.append(exclusion)
            
            if exclusions:
                raw_data["exclusions"] = exclusions
    
    except Exception as e:
        # If anything fails, return empty dict and let normal flow handle it
        pass
    
    return raw_data


def add_data_mapping_sheet(wb, json_file_path, bundle, logger):
    """Sheet 1: Clinical data mapping from source extraction to FHIR resources."""
    ws = wb.create_sheet("Data Mapping", 0)
    styles = setup_styles()
    
    # Try to find corresponding raw LLM file
    raw_llm_file = None
    base_name = os.path.basename(json_file_path).replace(".json", "")
    logs_dir = "logs"
    
    if os.path.exists(logs_dir):
        for file in os.listdir(logs_dir):
            if file.startswith(base_name) and file.endswith("_raw_llm.json"):
                raw_llm_file = os.path.join(logs_dir, file)
                break
    
    raw_data = {}
    if raw_llm_file and os.path.exists(raw_llm_file):
        try:
            with open(raw_llm_file, "r", encoding="utf-8") as f:
                raw_data = json.load(f)
        except Exception as e:
            if logger:
                logger.warning(f"Could not load raw LLM file: {str(e)}")
    
    # If no raw data found, reconstruct from FHIR bundle
    if not raw_data:
        raw_data = _reconstruct_raw_data_from_bundle(bundle)
    
    row = 1
    
    # Title
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1)
    title.value = "Clinical Data Mapping: Source → FHIR Resources"
    title.font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 25
    row = 3
    
    # Find Organization and InsurancePlan resources
    org_data = None
    plan_data = None
    for entry in bundle.get("entry", []):
        resource = entry.get("resource", {})
        if resource.get("resourceType") == "Organization":
            org_data = resource
        elif resource.get("resourceType") == "InsurancePlan":
            plan_data = resource
    
    # Create mapping entries
    mappings = []
    
    if raw_data:
        # Organization mappings
        if org_data:
            if raw_data.get("organization"):
                mappings.append({
                    "source_field": "organization",
                    "source_value": raw_data.get("organization", ""),
                    "fhir_resource": "Organization",
                    "fhir_path": "Organization.name",
                    "fhir_value": org_data.get("name", "")
                })
            
            if raw_data.get("insurer_id"):
                mappings.append({
                    "source_field": "insurer_id",
                    "source_value": raw_data.get("insurer_id", ""),
                    "fhir_resource": "Organization",
                    "fhir_path": "Organization.identifier[0].value",
                    "fhir_value": org_data.get("identifier", [{}])[0].get("value", "") if org_data.get("identifier") else ""
                })
            
            if raw_data.get("telecom", {}).get("phone"):
                mappings.append({
                    "source_field": "telecom.phone",
                    "source_value": raw_data.get("telecom", {}).get("phone", ""),
                    "fhir_resource": "Organization",
                    "fhir_path": "Organization.telecom[0].value",
                    "fhir_value": next((t.get("value", "") for t in org_data.get("telecom", []) if t.get("system") == "phone"), "")
                })
            
            if raw_data.get("telecom", {}).get("email"):
                mappings.append({
                    "source_field": "telecom.email",
                    "source_value": raw_data.get("telecom", {}).get("email", ""),
                    "fhir_resource": "Organization",
                    "fhir_path": "Organization.telecom[1].value",
                    "fhir_value": next((t.get("value", "") for t in org_data.get("telecom", []) if t.get("system") == "email"), "")
                })
        
        # InsurancePlan mappings
        if plan_data:
            if raw_data.get("uin"):
                mappings.append({
                    "source_field": "uin",
                    "source_value": raw_data.get("uin", ""),
                    "fhir_resource": "InsurancePlan",
                    "fhir_path": "InsurancePlan.identifier[0].value",
                    "fhir_value": plan_data.get("identifier", [{}])[0].get("value", "") if plan_data.get("identifier") else ""
                })
            
            if raw_data.get("plan_name"):
                mappings.append({
                    "source_field": "plan_name",
                    "source_value": raw_data.get("plan_name", ""),
                    "fhir_resource": "InsurancePlan",
                    "fhir_path": "InsurancePlan.name",
                    "fhir_value": plan_data.get("name", "")
                })
            
            if raw_data.get("sum_insured"):
                mappings.append({
                    "source_field": "sum_insured",
                    "source_value": raw_data.get("sum_insured", ""),
                    "fhir_resource": "InsurancePlan",
                    "fhir_path": "InsurancePlan.plan[0].generalCost[0].cost.value",
                    "fhir_value": str(plan_data.get("plan", [{}])[0].get("generalCost", [{}])[0].get("cost", {}).get("value", "")) if plan_data.get("plan") else ""
                })
            
            # Benefits mapping
            raw_benefits = raw_data.get("benefits", [])
            fhir_benefits = []
            for coverage in plan_data.get("coverage", []):
                for benefit in coverage.get("benefit", []):
                    fhir_benefits.append(benefit.get("type", {}).get("text", ""))
            
            for idx, raw_benefit in enumerate(raw_benefits[:3]):  # Limit to first 3 for clarity
                if idx < len(fhir_benefits):
                    mappings.append({
                        "source_field": f"benefits[{idx}].name",
                        "source_value": raw_benefit.get("name", ""),
                        "fhir_resource": "InsurancePlan",
                        "fhir_path": f"InsurancePlan.coverage[].benefit[{idx}].type.text",
                        "fhir_value": fhir_benefits[idx]
                    })
            
            # Exclusions mapping
            raw_exclusions = raw_data.get("exclusions", [])
            fhir_exclusions = []
            for ext in plan_data.get("extension", []):
                if ext.get("url") == "https://nrces.in/ndhm/fhir/r4/StructureDefinition/Claim-Exclusion":
                    for item in ext.get("extension", []):
                        if item.get("url") == "category":
                            fhir_exclusions.append(item.get("valueCodeableConcept", {}).get("text", ""))
                            break
            
            for idx, raw_excl in enumerate(raw_exclusions[:2]):  # Limit to first 2 for clarity
                if idx < len(fhir_exclusions):
                    mappings.append({
                        "source_field": f"exclusions[{idx}].name",
                        "source_value": raw_excl.get("name", ""),
                        "fhir_resource": "InsurancePlan",
                        "fhir_path": f"InsurancePlan.extension[{idx}].extension[0].valueCodeableConcept.text",
                        "fhir_value": fhir_exclusions[idx]
                    })
    
    # Write mappings to sheet
    format_header(ws, row, ["Source Field", "Source Value", "FHIR Resource", "FHIR Path", "Mapped Value"], styles)
    row += 1
    
    for mapping in mappings:
        ws.cell(row=row, column=1).value = mapping["source_field"]
        ws.cell(row=row, column=2).value = str(mapping["source_value"])[:100]  # Truncate long values
        ws.cell(row=row, column=3).value = mapping["fhir_resource"]
        ws.cell(row=row, column=4).value = mapping["fhir_path"]
        ws.cell(row=row, column=5).value = str(mapping["fhir_value"])[:100]  # Truncate long values
        
        for col in [1, 2, 3, 4, 5]:
            ws.cell(row=row, column=col).border = styles["border"]
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1
    
    if not mappings:
        ws.cell(row=row, column=1).value = "No raw data available for mapping"
    
    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 30


def add_mapping_sheet(wb, logger):
    """Sheet 2: Benefit mapping rules and configuration."""
    ws = wb.create_sheet("Mapping Rules", 1)
    styles = setup_styles()
    
    try:
        with open("config/mapping.yaml", "r") as f:
            mapping_config = yaml.safe_load(f)
    except FileNotFoundError:
        logger.warning("mapping.yaml not found")
        mapping_config = {}
    
    row = 1
    
    # Title
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1)
    title.value = "FHIR Mapping Configuration"
    title.font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 25
    row = 3
    
    # Benefit Mapping
    format_subheader(ws, row, "Benefit Mapping Rules", styles, 2)
    row += 1
    
    format_header(ws, row, ["Keyword (Source)", "Mapped To (FHIR Display Name)"], styles)
    row += 1
    
    benefit_mapping = mapping_config.get("benefit_mapping", {})
    for keyword, display_name in sorted(benefit_mapping.items()):
        ws.cell(row=row, column=1).value = keyword
        ws.cell(row=row, column=2).value = display_name
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = styles["border"]
        row += 1
    
    row += 1
    
    # Exclusion Categories
    format_subheader(ws, row, "Exclusion Categories", styles, 2)
    row += 1
    
    format_header(ws, row, ["Keyword", "Exclusion Type"], styles)
    row += 1
    
    exclusion_mapping = mapping_config.get("exclusion_categories", {})
    for keyword, excl_type in sorted(exclusion_mapping.items()):
        ws.cell(row=row, column=1).value = keyword
        ws.cell(row=row, column=2).value = excl_type
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = styles["border"]
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40


def add_organization_sheet(wb, bundle, logger):
    """Sheet 3: Organization resource details."""
    ws = wb.create_sheet("Organization", 2)
    styles = setup_styles()
    
    row = 1
    ws.merge_cells("A1:D1")
    title = ws.cell(row=1, column=1)
    title.value = "Organization Details"
    title.font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 25
    row = 3
    
    # Find Organization resource
    org_data = None
    for entry in bundle.get("entry", []):
        resource = entry.get("resource", {})
        if resource.get("resourceType") == "Organization":
            org_data = resource
            break
    
    if org_data:
        format_header(ws, row, ["Property", "Value"], styles)
        row += 1
        
        # Basic Info
        properties = [
            ("Resource Type", org_data.get("resourceType", "")),
            ("ID", org_data.get("id", "")),
            ("Name", org_data.get("name", "")),
            ("Active", org_data.get("active", "")),
        ]
        
        # Identifiers
        identifiers = org_data.get("identifier", [])
        for idx, ident in enumerate(identifiers):
            code = ident.get("identifier", {}).get("type", {}).get("coding", [{}])[0].get("code", "")
            value = ident.get("value", "")
            properties.append((f"Identifier {idx+1} (Code: {code})", value))
        
        # Telecom
        telecom_list = org_data.get("telecom", [])
        for idx, contact in enumerate(telecom_list):
            system = contact.get("system", "")
            value = contact.get("value", "")
            properties.append((f"Telecom {idx+1} ({system})", value))
        
        for prop_name, prop_value in properties:
            ws.cell(row=row, column=1).value = prop_name
            ws.cell(row=row, column=2).value = str(prop_value)
            for col in [1, 2]:
                ws.cell(row=row, column=col).border = styles["border"]
            row += 1
    else:
        ws.cell(row=row, column=1).value = "No Organization resource found"
    
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50


def add_insurance_plan_sheet(wb, bundle, logger):
    """Sheet 4: InsurancePlan resource details."""
    ws = wb.create_sheet("Insurance Plan", 3)
    styles = setup_styles()
    
    row = 1
    ws.merge_cells("A1:D1")
    title = ws.cell(row=1, column=1)
    title.value = "Insurance Plan Details"
    title.font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 25
    row = 3
    
    # Find InsurancePlan resource
    plan_data = None
    for entry in bundle.get("entry", []):
        resource = entry.get("resource", {})
        if resource.get("resourceType") == "InsurancePlan":
            plan_data = resource
            break
    
    if plan_data:
        # Basic Info
        format_subheader(ws, row, "Basic Information", styles, 2)
        row += 1
        
        format_header(ws, row, ["Property", "Value"], styles)
        row += 1
        
        properties = [
            ("Plan Name", plan_data.get("name", "")),
            ("Status", plan_data.get("status", "")),
            ("Period Start", plan_data.get("period", {}).get("start", "")),
            ("Period End", plan_data.get("period", {}).get("end", "")),
            ("Plan Type", plan_data.get("plan", [{}])[0].get("type", {}).get("coding", [{}])[0].get("display", "") if plan_data.get("plan") else ""),
        ]
        
        for prop_name, prop_value in properties:
            ws.cell(row=row, column=1).value = prop_name
            ws.cell(row=row, column=2).value = str(prop_value)
            for col in [1, 2]:
                ws.cell(row=row, column=col).border = styles["border"]
            row += 1
        
        row += 1
        
        # Coverage and Benefits
        format_subheader(ws, row, "Coverage & Benefits", styles, 4)
        row += 1
        
        format_header(ws, row, ["Coverage Type", "Benefit Name", "Benefit Code", "Display"], styles)
        row += 1
        
        coverage_list = plan_data.get("coverage", [])
        for cov in coverage_list:
            cov_type = cov.get("type", {}).get("coding", [{}])[0].get("display", "")
            benefits = cov.get("benefit", [])
            
            for benefit in benefits:
                benefit_name = benefit.get("type", {}).get("text", "")
                benefit_code = benefit.get("type", {}).get("coding", [{}])[0].get("code", "")
                benefit_display = benefit.get("type", {}).get("coding", [{}])[0].get("display", "")
                
                ws.cell(row=row, column=1).value = cov_type
                ws.cell(row=row, column=2).value = benefit_name
                ws.cell(row=row, column=3).value = benefit_code
                ws.cell(row=row, column=4).value = benefit_display
                for col in [1, 2, 3, 4]:
                    ws.cell(row=row, column=col).border = styles["border"]
                row += 1
    else:
        ws.cell(row=row, column=1).value = "No InsurancePlan resource found"
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35


def add_exclusions_sheet(wb, bundle, logger):
    """Sheet 5: Claim Exclusions."""
    ws = wb.create_sheet("Exclusions", 4)
    styles = setup_styles()
    
    row = 1
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1)
    title.value = "Claim Exclusions"
    title.font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 25
    row = 3
    
    # Find InsurancePlan resource
    plan_data = None
    for entry in bundle.get("entry", []):
        resource = entry.get("resource", {})
        if resource.get("resourceType") == "InsurancePlan":
            plan_data = resource
            break
    
    if plan_data:
        format_header(ws, row, ["Exclusion Category", "Statement"], styles)
        row += 1
        
        extensions = plan_data.get("extension", [])
        for ext in extensions:
            if ext.get("url") == "https://nrces.in/ndhm/fhir/r4/StructureDefinition/Claim-Exclusion":
                ext_items = ext.get("extension", [])
                category = ""
                statement = ""
                
                for item in ext_items:
                    if item.get("url") == "category":
                        category = item.get("valueCodeableConcept", {}).get("text", "")
                    elif item.get("url") == "statement":
                        statement = item.get("valueString", "")
                
                if category or statement:
                    ws.cell(row=row, column=1).value = category
                    ws.cell(row=row, column=2).value = statement
                    for col in [1, 2]:
                        ws.cell(row=row, column=col).border = styles["border"]
                        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                    row += 1
    else:
        ws.cell(row=row, column=1).value = "No InsurancePlan resource found"
    
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 70
    ws.row_dimensions_height = 20


def generate_excel_from_json(json_file_path, output_excel_path=None, logger_obj=None):
    """
    Generate Excel file from FHIR JSON bundle.
    
    Args:
        json_file_path: Path to the JSON bundle file
        output_excel_path: Path where to save the Excel file (default: same name with .xlsx)
        logger_obj: Logger object for logging
    
    Returns:
        Path to the generated Excel file
    """
    global logger
    logger = logger_obj
    
    if not os.path.exists(json_file_path):
        if logger:
            logger.error(f"JSON file not found: {json_file_path}")
        return None
    
    # Load JSON bundle
    with open(json_file_path, "r", encoding="utf-8") as f:
        bundle = json.load(f)
    
    if logger:
        logger.info(f"Loaded FHIR bundle from {json_file_path}")
    
    # Determine output path
    if output_excel_path is None:
        output_excel_path = json_file_path.replace(".json", "_mapping.xlsx")
    
    # Create workbook and add sheets
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    try:
        add_data_mapping_sheet(wb, json_file_path, bundle, logger)
        add_mapping_sheet(wb, logger)
        add_organization_sheet(wb, bundle, logger)
        add_insurance_plan_sheet(wb, bundle, logger)
        add_exclusions_sheet(wb, bundle, logger)
        
        # Save workbook
        wb.save(output_excel_path)
        
        if logger:
            logger.info(f"Excel mapping file created: {output_excel_path}")
        
        return output_excel_path
    
    except Exception as e:
        if logger:
            logger.error(f"Error generating Excel file: {str(e)}")
        raise


def process_all_outputs(output_dir="output", logger_obj=None):
    """
    Process all JSON files in output directory and generate Excel files.
    
    Args:
        output_dir: Directory containing JSON output files
        logger_obj: Logger object for logging
    
    Returns:
        List of generated Excel file paths
    """
    global logger
    logger = logger_obj
    
    generated_files = []
    
    if not os.path.exists(output_dir):
        if logger:
            logger.warning(f"Output directory not found: {output_dir}")
        return generated_files
    
    # Search for JSON files in output directory
    for root, dirs, files in os.walk(output_dir):
        for file in files:
            if file.endswith(".json"):
                json_path = os.path.join(root, file)
                try:
                    excel_path = generate_excel_from_json(json_path, logger_obj=logger_obj)
                    if excel_path:
                        generated_files.append(excel_path)
                except Exception as e:
                    if logger:
                        logger.error(f"Failed to process {json_path}: {str(e)}")
    
    return generated_files
